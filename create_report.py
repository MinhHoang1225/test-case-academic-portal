import os
import sys
import glob
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def process_single_module_workbook(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Find data sheet (non-Report)
    data_ws = None
    data_sheet_name = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() != "report":
            data_ws = wb[sheet_name]
            data_sheet_name = sheet_name
            break
            
    if not data_ws:
        print(f"No data sheet found in {xlsx_path}")
        return

    max_r = data_ws.max_row
    rows = list(data_ws.iter_rows(values_only=True))
    if not rows:
        return
        
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    
    p_idx = headers.index("Priority") if "Priority" in headers else -1
    fid_idx = headers.index("Feature ID") if "Feature ID" in headers else -1
    fname_idx = headers.index("Feature Name") if "Feature Name" in headers else -1

    p_col_letter = get_column_letter(p_idx + 1) if p_idx != -1 else "K"
    fid_col_letter = get_column_letter(fid_idx + 1) if fid_idx != -1 else "C"

    # Unique features in ordering
    unique_features = []
    seen = set()
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        fid = str(row[fid_idx]).strip() if fid_idx != -1 and fid_idx < len(row) and row[fid_idx] is not None else "GEN"
        fname = str(row[fname_idx]).strip() if fname_idx != -1 and fname_idx < len(row) and row[fname_idx] is not None else "General"
        
        fkey = (fid, fname)
        if fkey not in seen:
            seen.add(fkey)
            unique_features.append(fkey)

    # Remove existing Report sheet if present
    if "Report" in wb.sheetnames:
        del wb["Report"]

    ws = wb.create_sheet(title="Report", index=0)
    ws.views.sheetView[0].showGridLines = True

    # Styling
    FONT_TITLE = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="555555")
    FONT_CARD_LABEL = Font(name="Segoe UI", size=9, bold=True, color="64748B")
    FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    FONT_BODY = Font(name="Segoe UI", size=10, color="000000")
    FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color="000000")

    FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    FILL_TOTAL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    BORDER_THIN = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    BORDER_TOTAL = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color="1E293B"),
        bottom=Side(style="double", color="1E293B")
    )

    base_name = os.path.basename(xlsx_path).replace("_test_cases.xlsx", "").replace(".xlsx", "")
    mod_name = base_name.replace("module", "Module ").title()
    ws.cell(row=1, column=1, value=f"{mod_name} - Feature & Priority Summary Report").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Summary of test cases per feature, calculated via Excel formulas").font = FONT_SUBTITLE

    # Table layout
    start_row = 7
    header_row = start_row + 1
    table_headers = ["Feature ID", "Feature Name", "High Priority", "Medium Priority", "Low Priority", "Total TCs"]
    ws.row_dimensions[header_row].height = 26

    for col_i, h in enumerate(table_headers, start=1):
        cell = ws.cell(row=header_row, column=col_i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center" if col_i in [1, 3, 4, 5, 6] else "left", vertical="center")
        cell.border = BORDER_THIN

    row_idx = header_row + 1
    for fid, fname in unique_features:
        ws.row_dimensions[row_idx].height = 22
        
        # EXCEL FORMULAS FOR COUNTING HIGH, MEDIUM, LOW, TOTAL
        f_high = f'=COUNTIFS(\'{data_sheet_name}\'!${fid_col_letter}$2:${fid_col_letter}${max_r}, A{row_idx}, \'{data_sheet_name}\'!${p_col_letter}$2:${p_col_letter}${max_r}, "High")'
        f_med  = f'=COUNTIFS(\'{data_sheet_name}\'!${fid_col_letter}$2:${fid_col_letter}${max_r}, A{row_idx}, \'{data_sheet_name}\'!${p_col_letter}$2:${p_col_letter}${max_r}, "Medium")'
        f_low  = f'=COUNTIFS(\'{data_sheet_name}\'!${fid_col_letter}$2:${fid_col_letter}${max_r}, A{row_idx}, \'{data_sheet_name}\'!${p_col_letter}$2:${p_col_letter}${max_r}, "Low")'
        f_tot  = f'=SUM(C{row_idx}:E{row_idx})'

        vals = [fid, fname, f_high, f_med, f_low, f_tot]
        is_even = (row_idx % 2 == 0)
        row_fill = FILL_ZEBRA if not is_even else PatternFill(fill_type=None)

        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_i, value=val)
            cell.font = FONT_BODY
            if col_i == 1:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
            cell.alignment = Alignment(horizontal="center" if col_i in [1, 3, 4, 5, 6] else "left", vertical="center")
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = BORDER_THIN

            if col_i == 3:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="A50E0E")
            elif col_i == 4:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="B06000")
            elif col_i == 5:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="137333")

        row_idx += 1

    # Total Row Formula
    tot_row = row_idx
    ws.row_dimensions[tot_row].height = 24
    tot_vals = [
        "Total",
        f'=COUNTA(A9:A{tot_row-1})&" Features"',
        f'=SUM(C9:C{tot_row-1})',
        f'=SUM(D9:D{tot_row-1})',
        f'=SUM(E9:E{tot_row-1})',
        f'=SUM(F9:F{tot_row-1})'
    ]
    for col_i, val in enumerate(tot_vals, start=1):
        cell = ws.cell(row=tot_row, column=col_i, value=val)
        cell.font = FONT_TOTAL
        cell.alignment = Alignment(horizontal="center" if col_i in [1, 3, 4, 5, 6] else "left", vertical="center")
        cell.fill = FILL_TOTAL
        cell.border = BORDER_TOTAL

    # Metric KPI Cards
    cards = [
        ("Total Features", f'=COUNTA(A9:A{tot_row-1})', "1E293B", "F1F5F9"),
        ("Total Test Cases", f'=F{tot_row}', "1F4E78", "E2E8F0"),
        ("High Priority Features", f'=COUNTIF(C9:C{tot_row-1}, ">0")', "A50E0E", "FCE8E6"),
        ("Medium Priority Features", f'=COUNTIFS(C9:C{tot_row-1}, 0, D9:D{tot_row-1}, ">0")', "B06000", "FEF7E0"),
        ("Low Priority Features", f'=COUNTIFS(C9:C{tot_row-1}, 0, D9:D{tot_row-1}, 0, E9:E{tot_row-1}, ">0")', "137333", "E6F4EA"),
    ]

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 32

    for col_i, (label, val, fg_color, bg_color) in enumerate(cards, start=1):
        c_lbl = ws.cell(row=4, column=col_i, value=label)
        c_lbl.font = FONT_CARD_LABEL
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")

        c_val = ws.cell(row=5, column=col_i, value=val)
        c_val.font = Font(name="Segoe UI", size=18, bold=True, color=fg_color)
        c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = BORDER_THIN

    # Table section title
    ws.cell(row=start_row, column=1, value="Feature Categorization & Priority Breakdown").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E78")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = -20
    chart.title = "Test Cases by Feature (Grouped by Priority)"
    chart.style = 10
    chart.height = 14
    chart.width = 22
    chart.y_axis.title = "Number of Test Cases"
    chart.x_axis.title = "Feature ID"

    data = Reference(ws, min_col=3, min_row=header_row, max_col=5, max_row=tot_row - 1)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=tot_row - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    colors = ["FF0000", "FFFF00", "00FF00"]
    for idx, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = colors[idx]

    ws.add_chart(chart, "H4")

    wb.save(xlsx_path)
    print(f"Successfully generated Report sheet with Excel Formulas & Bar Chart in {xlsx_path}")

def process_master_report(tests_excel_dir):
    master_xlsx_path = os.path.join(tests_excel_dir, "academic_portal_all_test_cases.xlsx")
    
    if os.path.exists(master_xlsx_path):
        wb = openpyxl.load_workbook(master_xlsx_path)
    else:
        wb = openpyxl.Workbook()

    # Get all module sheets
    module_sheet_names = [s for s in wb.sheetnames if s.lower() not in ["report", "summary"]]
    module_sheet_names.sort()

    if "Report" in wb.sheetnames:
        del wb["Report"]

    ws = wb.create_sheet(title="Report", index=0)
    ws.views.sheetView[0].showGridLines = True

    FONT_TITLE = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="555555")
    FONT_CARD_LABEL = Font(name="Segoe UI", size=9, bold=True, color="64748B")
    FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    FONT_BODY = Font(name="Segoe UI", size=10, color="000000")
    FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color="000000")

    FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    FILL_TOTAL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    BORDER_THIN = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    BORDER_TOTAL = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color="1E293B"),
        bottom=Side(style="double", color="1E293B")
    )

    ws.cell(row=1, column=1, value="Academic Portal - Test Coverage & Feature Report").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Summary of test cases per module, calculated via Excel formulas").font = FONT_SUBTITLE

    start_row = 7
    header_row = start_row + 1
    headers = ["Module Name", "High Priority", "Medium Priority", "Low Priority", "Total TCs"]
    ws.row_dimensions[header_row].height = 26

    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")
        cell.border = BORDER_THIN

    row_idx = header_row + 1
    for m_sheet in module_sheet_names:
        ws.row_dimensions[row_idx].height = 22
        
        # Dynamically determine Priority column letter in target module worksheet
        m_ws = wb[m_sheet]
        m_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in m_ws[1]]
        p_idx = m_headers.index("Priority") if "Priority" in m_headers else -1
        p_col_letter = get_column_letter(p_idx + 1) if p_idx != -1 else "M"

        f_high = f'=COUNTIF(\'{m_sheet}\'!${p_col_letter}$2:${p_col_letter}$500, "High")'
        f_med  = f'=COUNTIF(\'{m_sheet}\'!${p_col_letter}$2:${p_col_letter}$500, "Medium")'
        f_low  = f'=COUNTIF(\'{m_sheet}\'!${p_col_letter}$2:${p_col_letter}$500, "Low")'
        f_tot  = f'=SUM(B{row_idx}:D{row_idx})'

        vals = [m_sheet, f_high, f_med, f_low, f_tot]
        is_even = (row_idx % 2 == 0)
        row_fill = FILL_ZEBRA if not is_even else PatternFill(fill_type=None)

        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_i, value=val)
            cell.font = FONT_BODY
            cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = BORDER_THIN

            if col_i == 2:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="A50E0E")
            elif col_i == 3:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="B06000")
            elif col_i == 4:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="137333")

        row_idx += 1

    tot_row = row_idx
    ws.row_dimensions[tot_row].height = 24
    tot_vals = [
        "Total",
        f'=SUM(B9:B{tot_row-1})',
        f'=SUM(C9:C{tot_row-1})',
        f'=SUM(D9:D{tot_row-1})',
        f'=SUM(E9:E{tot_row-1})'
    ]
    for col_i, val in enumerate(tot_vals, start=1):
        cell = ws.cell(row=tot_row, column=col_i, value=val)
        cell.font = FONT_TOTAL
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")
        cell.fill = FILL_TOTAL
        cell.border = BORDER_TOTAL

    # Metric KPI Cards
    cards = [
        ("Total Modules", f'=COUNTA(A9:A{tot_row-1})', "1E293B", "F1F5F9"),
        ("Total Test Cases", f'=E{tot_row}', "1F4E78", "E2E8F0"),
        ("High Priority TCs", f'=B{tot_row}', "A50E0E", "FCE8E6"),
        ("Medium Priority TCs", f'=C{tot_row}', "B06000", "FEF7E0"),
        ("Low Priority TCs", f'=D{tot_row}', "137333", "E6F4EA"),
    ]

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 32

    for col_i, (label, val, fg_color, bg_color) in enumerate(cards, start=1):
        c_lbl = ws.cell(row=4, column=col_i, value=label)
        c_lbl.font = FONT_CARD_LABEL
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")

        c_val = ws.cell(row=5, column=col_i, value=val)
        c_val.font = Font(name="Segoe UI", size=18, bold=True, color=fg_color)
        c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = BORDER_THIN

    # Table section title
    ws.cell(row=start_row, column=1, value="Module Test Case Breakdown by Priority").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E78")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = -20
    chart.title = "Test Cases by Module (Grouped by Priority)"
    chart.style = 10
    chart.height = 14
    chart.width = 22
    chart.y_axis.title = "Number of Test Cases"
    chart.x_axis.title = "Module"

    data = Reference(ws, min_col=2, min_row=header_row, max_col=4, max_row=tot_row - 1)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=tot_row - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    colors = ["FF0000", "FFFF00", "00FF00"]
    for idx, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = colors[idx]

    ws.add_chart(chart, "G4")

    wb.save(master_xlsx_path)
    print(f"Successfully generated Master Report sheet with Excel Formulas & Bar Chart in {master_xlsx_path}")

def main():
    base_tests_dir = os.path.join(os.getcwd(), "tests")
    excel_dir = os.path.join(base_tests_dir, "excel") if os.path.isdir(os.path.join(base_tests_dir, "excel")) else base_tests_dir

    if len(sys.argv) > 1:
        target_path = sys.argv[1]
        if not os.path.isabs(target_path):
            target_path = os.path.join(os.getcwd(), target_path)
        if os.path.isfile(target_path):
            if "academic_portal_all" in target_path:
                process_master_report(os.path.dirname(target_path))
            else:
                process_single_module_workbook(target_path)
        else:
            print(f"File not found: {target_path}")
    else:
        # Process Master Report
        process_master_report(excel_dir)

        # Process all individual module XLSX files
        module_files = sorted(glob.glob(os.path.join(excel_dir, "module*_test_cases.xlsx")))
        for m_file in module_files:
            process_single_module_workbook(m_file)

if __name__ == "__main__":
    main()
