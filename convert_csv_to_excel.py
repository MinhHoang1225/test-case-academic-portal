import os
import glob
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colors
HEADER_BG = "1F4E78"        # Navy blue
HEADER_FG = "FFFFFF"        # White text
ZEBRA_BG  = "F8FAFC"        # Light slate/blue tint
BORDER_COLOR = "CBD5E1"     # Light gray border

PRIORITY_MAP = {
    "HIGH": "High",
    "CAO": "High",
    "MEDIUM": "Medium",
    "TRUNG BÌNH": "Medium",
    "TRUNG BINH": "Medium",
    "LOW": "Low",
    "THẤP": "Low",
    "THAP": "Low"
}

PRIORITY_STYLES = {
    "High": {
        "fill": PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="A50E0E")
    },
    "Medium": {
        "fill": PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="B06000")
    },
    "Low": {
        "fill": PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="137333")
    }
}

# Column widths
COLUMN_WIDTH_MAP = {
    "ID": 15,
    "Module": 14,
    "Feature ID": 14,
    "Feature Name": 28,
    "Title": 32,
    "Objective": 36,
    "Preconditions": 30,
    "TestData": 28,
    "Steps": 48,
    "ExpectedResult": 48,
    "Actual Result": 36,
    "Status": 16,
    "Priority": 14,
    "Notes": 32
}

def clean_cell_text(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s

def normalize_priority(val):
    v = str(val).strip().upper()
    return PRIORITY_MAP.get(v, str(val).strip().capitalize())

def format_worksheet(ws, headers, rows):
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Segoe UI", size=10, color="000000")
    thin_side = Side(border_style="thin", color=BORDER_COLOR)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Write Headers
    ws.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=h_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

        width = COLUMN_WIDTH_MAP.get(h_name, 25)
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    # Write Rows
    for row_idx, row_data in enumerate(rows, start=2):
        is_even = (row_idx % 2 == 0)
        row_fill = PatternFill(start_color=ZEBRA_BG, end_color=ZEBRA_BG, fill_type="solid") if not is_even else PatternFill(fill_type=None)
        
        max_lines = 1

        for col_idx, cell_value in enumerate(row_data, start=1):
            if col_idx > len(headers):
                break
            h_name = headers[col_idx-1]
            text = clean_cell_text(cell_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            
            cell.font = body_font
            cell.border = cell_border

            h_align = "left"
            if h_name in ["ID", "Module", "Feature ID", "Priority", "Status"]:
                h_align = "center"

            cell.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=True)

            if h_name == "Priority":
                p_norm = normalize_priority(text)
                if p_norm in PRIORITY_STYLES:
                    cell.fill = PRIORITY_STYLES[p_norm]["fill"]
                    cell.font = PRIORITY_STYLES[p_norm]["font"]
                    cell.value = p_norm
            elif h_name == "Status":
                s_lower = text.lower()
                if s_lower in ["passed", "pass"]:
                    cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="137333")
                    cell.value = "Passed"
                elif s_lower in ["failed", "fail"]:
                    cell.fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="A50E0E")
                    cell.value = "Failed"
                else:
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="64748B")
                    cell.value = "Not Run"
            elif h_name == "Feature ID":
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
                if not is_even:
                    cell.fill = row_fill
            else:
                if not is_even:
                    cell.fill = row_fill

            lines = text.count('\n') + 1
            if len(text) > 40:
                lines += len(text) // 40
            if lines > max_lines:
                max_lines = lines

        calc_height = max(22, min(180, max_lines * 15))
        ws.row_dimensions[row_idx].height = calc_height

    # Add Data Validation Dropdowns for Status and Priority (Title Case)
    max_row_idx = len(rows) + 1
    if "Status" in headers and max_row_idx >= 2:
        status_col_letter = get_column_letter(headers.index("Status") + 1)
        dv_status = DataValidation(type="list", formula1='"Passed,Failed,Not Run"', allow_blank=True)
        dv_status.error = 'Please select a valid state: Passed, Failed, or Not Run'
        dv_status.errorTitle = 'Invalid Status'
        ws.add_data_validation(dv_status)
        dv_status.add(f"{status_col_letter}2:{status_col_letter}{max_row_idx+100}")

    if "Priority" in headers and max_row_idx >= 2:
        priority_col_letter = get_column_letter(headers.index("Priority") + 1)
        dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
        dv_priority.error = 'Please select a valid priority: High, Medium, or Low'
        dv_priority.errorTitle = 'Invalid Priority'
        ws.add_data_validation(dv_priority)
        dv_priority.add(f"{priority_col_letter}2:{priority_col_letter}{max_row_idx+100}")

    ws.auto_filter.ref = ws.dimensions

def read_csv_file(csv_path, xlsx_path=None):
    rows = []
    headers = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if not row:
                continue
            if i == 0:
                headers = [h.strip().strip('"').replace('\ufeff', '') for h in row]
                headers = ["Preconditions" if h == "Precondition" else h for h in headers]
            else:
                rows.append(row)

    # Load existing Excel data if available to preserve Actual Result & Status
    existing_data = {}
    if xlsx_path and os.path.exists(xlsx_path):
        try:
            with open(xlsx_path, 'rb') as f_ex:
                wb_ex = openpyxl.load_workbook(f_ex, data_only=True)
                for sheetname in wb_ex.sheetnames:
                    if sheetname.lower() != "report":
                        ws_ex = wb_ex[sheetname]
                        ex_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_ex[1]]
                        id_i = ex_headers.index("ID") if "ID" in ex_headers else 0
                        act_i = ex_headers.index("Actual Result") if "Actual Result" in ex_headers else -1
                        st_i = ex_headers.index("Status") if "Status" in ex_headers else -1

                        for r_cells in ws_ex.iter_rows(min_row=2, values_only=True):
                            if r_cells and r_cells[id_i]:
                                t_id = str(r_cells[id_i]).strip()
                                act_val = r_cells[act_i] if act_i != -1 and act_i < len(r_cells) else None
                                st_val = r_cells[st_i] if st_i != -1 and st_i < len(r_cells) else None
                                existing_data[t_id] = (act_val, st_val)
                        break
                wb_ex.close()
        except Exception as e:
            print(f"  Note: Could not read existing excel {xlsx_path}: {e}")

    # Ensure Actual Result and Status fields are added to headers if missing
    if "Actual Result" not in headers or "Status" not in headers:
        new_headers = []
        for h in headers:
            new_headers.append(h)
            if h == "ExpectedResult":
                if "Actual Result" not in headers:
                    new_headers.append("Actual Result")
                if "Status" not in headers:
                    new_headers.append("Status")

        exp_idx = headers.index("ExpectedResult") if "ExpectedResult" in headers else -1
        id_idx = headers.index("ID") if "ID" in headers else 0

        new_rows = []
        for r in rows:
            r_new = list(r)
            tc_id = r[id_idx].strip() if id_idx < len(r) and r[id_idx] else ""
            ex_act, ex_st = existing_data.get(tc_id, (None, None))

            if exp_idx != -1 and exp_idx < len(r_new):
                insert_idx = exp_idx + 1
                if "Actual Result" not in headers:
                    act_to_insert = ex_act if ex_act is not None else ""
                    r_new.insert(insert_idx, act_to_insert)
                    insert_idx += 1
                if "Status" not in headers:
                    st_to_insert = ex_st if ex_st is not None else "Not Run"
                    r_new.insert(insert_idx, st_to_insert)
            new_rows.append(r_new)

        headers = new_headers
        rows = new_rows

    return headers, rows

def create_summary_sheet(wb, module_data):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="Academic Portal - Comprehensive Test Suite & Feature Breakdown")
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    total_test_cases = 0
    total_high = 0
    total_medium = 0
    total_low = 0
    module_summary_rows = []
    feature_summary_rows = []

    for mod_num, (mod_title, headers, rows) in enumerate(module_data, start=1):
        mod_tc_count = len(rows)
        p_col_idx = headers.index("Priority") if "Priority" in headers else -1
        fid_col_idx = headers.index("Feature ID") if "Feature ID" in headers else -1
        fname_col_idx = headers.index("Feature Name") if "Feature Name" in headers else -1

        high_cnt = 0
        med_cnt = 0
        low_cnt = 0
        feature_stats = {}

        for r in rows:
            p_norm = normalize_priority(r[p_col_idx]) if p_col_idx != -1 and p_col_idx < len(r) else None
            if p_norm == "High":
                high_cnt += 1
            elif p_norm == "Medium":
                med_cnt += 1
            elif p_norm == "Low":
                low_cnt += 1

            if fid_col_idx != -1 and fname_col_idx != -1 and fid_col_idx < len(r) and fname_col_idx < len(r):
                fid_val = r[fid_col_idx]
                fname_val = r[fname_col_idx]
                fkey = (mod_title, fid_val, fname_val)
                if fkey not in feature_stats:
                    feature_stats[fkey] = {"count": 0, "high": 0, "medium": 0, "low": 0}
                feature_stats[fkey]["count"] += 1
                if p_norm == "High":
                    feature_stats[fkey]["high"] += 1
                elif p_norm == "Medium":
                    feature_stats[fkey]["medium"] += 1
                elif p_norm == "Low":
                    feature_stats[fkey]["low"] += 1

        total_test_cases += mod_tc_count
        total_high += high_cnt
        total_medium += med_cnt
        total_low += low_cnt

        module_summary_rows.append((mod_title, mod_tc_count, high_cnt, med_cnt, low_cnt))
        for (mname, fid, fname), fstat in sorted(feature_stats.items(), key=lambda x: x[0][1]):
            feature_summary_rows.append((mname, fid, fname, fstat["count"], fstat["high"], fstat["medium"], fstat["low"]))

    # Metric Cards Row
    metrics = [
        ("Total Modules", len(module_data), "1E293B", "F1F5F9"),
        ("Total Features", len(feature_summary_rows), "1F4E78", "E2E8F0"),
        ("Total Test Cases", total_test_cases, "1F4E78", "E2E8F0"),
        ("High Priority", total_high, "A50E0E", "FCE8E6"),
        ("Medium Priority", total_medium, "B06000", "FEF7E0"),
    ]

    for idx, (label, val, fg_color, bg_color) in enumerate(metrics):
        l_cell = ws.cell(row=3, column=idx+1, value=label)
        l_cell.font = Font(name="Segoe UI", size=9, bold=True, color="555555")
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        v_cell = ws.cell(row=4, column=idx+1, value=val)
        v_cell.font = Font(name="Segoe UI", size=18, bold=True, color=fg_color)
        v_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        thin_side = Side(border_style="thin", color="CBD5E1")
        v_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 35

    # Section 1: Module Level Breakdown (row 6)
    ws.cell(row=6, column=1, value="1. Module Summary Overview").font = Font(name="Segoe UI", size=13, bold=True, color="1F4E78")
    table_headers = ["Module Name", "Total Test Cases", "High Priority", "Medium Priority", "Low Priority"]
    ws.row_dimensions[7].height = 26
    for col_idx, th_name in enumerate(table_headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=th_name)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = Border(left=Side(style="thin", color="1E293B"), right=Side(style="thin", color="1E293B"), top=Side(style="thin", color="1E293B"), bottom=Side(style="thin", color="1E293B"))

    thin_border = Border(left=Side(style="thin", color="CBD5E1"),
                         right=Side(style="thin", color="CBD5E1"),
                         top=Side(style="thin", color="CBD5E1"),
                         bottom=Side(style="thin", color="CBD5E1"))

    for row_offset, (m_name, count, h_cnt, m_cnt, l_cnt) in enumerate(module_summary_rows, start=8):
        ws.row_dimensions[row_offset].height = 22
        vals = [m_name, count, h_cnt, m_cnt, l_cnt]
        is_even = (row_offset % 2 == 0)
        bg = "F8FAFC" if not is_even else "FFFFFF"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
            cell.fill = fill
            cell.border = thin_border

    tot_row_idx = 8 + len(module_summary_rows)
    ws.row_dimensions[tot_row_idx].height = 24
    tot_vals = ["Total", total_test_cases, total_high, total_medium, total_low]
    tot_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    tot_font = Font(name="Segoe UI", size=10, bold=True, color="000000")

    for col_idx, val in enumerate(tot_vals, start=1):
        cell = ws.cell(row=tot_row_idx, column=col_idx, value=val)
        cell.font = tot_font
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
        cell.fill = tot_fill
        cell.border = Border(top=Side(style="medium", color="1E293B"), bottom=Side(style="double", color="1E293B"), left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"))

    # Section 2: Feature Categorization Breakdown (row tot_row_idx + 3)
    feat_sec_start = tot_row_idx + 3
    ws.cell(row=feat_sec_start, column=1, value="2. Feature Categorization Breakdown").font = Font(name="Segoe UI", size=13, bold=True, color="1F4E78")
    
    f_headers = ["Module", "Feature ID", "Feature Name", "Test Cases", "High Priority", "Medium Priority", "Low Priority"]
    f_header_row = feat_sec_start + 1
    ws.row_dimensions[f_header_row].height = 26
    
    for col_idx, th_name in enumerate(f_headers, start=1):
        cell = ws.cell(row=f_header_row, column=col_idx, value=th_name)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx in [2,4,5,6,7] else "left", vertical="center")
        cell.border = Border(left=Side(style="thin", color="1E293B"), right=Side(style="thin", color="1E293B"), top=Side(style="thin", color="1E293B"), bottom=Side(style="thin", color="1E293B"))

    for row_offset, (mname, fid, fname, fcnt, fh, fm, fl) in enumerate(feature_summary_rows, start=f_header_row+1):
        ws.row_dimensions[row_offset].height = 22
        vals = [mname, fid, fname, fcnt, fh, fm, fl]
        is_even = (row_offset % 2 == 0)
        bg = "F8FAFC" if not is_even else "FFFFFF"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            if col_idx == 2:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
            cell.alignment = Alignment(horizontal="center" if col_idx in [2,4,5,6,7] else "left", vertical="center")
            cell.fill = fill
            cell.border = thin_border

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16

def process_csv_files(tests_dir):
    csv_dir = os.path.join(tests_dir, "csv") if os.path.isdir(os.path.join(tests_dir, "csv")) else tests_dir
    excel_dir = os.path.join(tests_dir, "excel") if os.path.isdir(os.path.join(tests_dir, "excel")) else tests_dir

    os.makedirs(excel_dir, exist_ok=True)

    csv_files = sorted(glob.glob(os.path.join(csv_dir, "*.csv")))
    print(f"Found {len(csv_files)} CSV files in {csv_dir}:")

    master_wb = openpyxl.Workbook()
    master_wb.remove(master_wb.active)

    module_data = []

    for csv_path in csv_files:
        base_name = os.path.basename(csv_path)
        name_without_ext = os.path.splitext(base_name)[0]
        xlsx_path = os.path.join(excel_dir, f"{name_without_ext}.xlsx")

        headers, rows = read_csv_file(csv_path, xlsx_path)
        print(f"  - Converting {base_name} ({len(rows)} test cases) -> {os.path.join('tests/excel', os.path.basename(xlsx_path))}")

        indiv_wb = openpyxl.Workbook()
        indiv_ws = indiv_wb.active
        sheet_title = name_without_ext.replace("_test_cases", "").capitalize()
        if len(sheet_title) > 30:
            sheet_title = sheet_title[:30]
        indiv_ws.title = sheet_title

        format_worksheet(indiv_ws, headers, rows)
        indiv_wb.save(xlsx_path)

        m_title = name_without_ext
        if "module" in m_title.lower():
            mod_part = m_title.lower().split("_")[0]
            mod_num = mod_part.replace("module", "")
            m_title = f"Module {mod_num}"
        else:
            m_title = name_without_ext[:30]

        master_ws = master_wb.create_sheet(title=m_title)
        format_worksheet(master_ws, headers, rows)

        module_data.append((m_title, headers, rows))

    create_summary_sheet(master_wb, module_data)

    master_xlsx_path = os.path.join(excel_dir, "academic_portal_all_test_cases.xlsx")
    master_wb.save(master_xlsx_path)
    print(f"\nCreated consolidated Master Workbook: {master_xlsx_path}")

    # Generate Report sheet and Bar Chart
    import create_report
    create_report.main()

if __name__ == "__main__":
    tests_directory = os.path.join(os.getcwd(), "tests")
    process_csv_files(tests_directory)
